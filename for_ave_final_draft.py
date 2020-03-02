# -*- coding: utf-8 -*-
"""
Created on Mon Jan 14 18:05:28 2019
@author: Simon
"""
import matplotlib.pyplot as plt
from datetime import date
from tqdm import tqdm
import numpy as np
import openpyxl
import os
import cv2
    
today = date.today()
plt.rcParams['figure.figsize'] = [10, 10]
script_dir = os.path.dirname(__file__)
images = 'images'
folder = 'output'
home = os.path.join(script_dir, images)
output = os.path.join(script_dir, folder)


def minimum_value_index(list):
    temp = []
    for i in list:
        add = np.sum(i)
        temp.append(add)
    marked = min(temp)
    return temp.index(marked)


def all_that_apply(list):
    temp = []
    checked = []
    indices = []
    for i in list:
        add = np.sum(i)
        temp.append(add)
    avg = (sum(temp) / len(temp)) - 300.0
    for i in temp:
        if i < avg:
            checked.append(i)
    for i in checked:
        indices.append(temp.index(i))
    return indices
    
    
def yes_no_q(yes, no):
    list = [yes, no]
    return minimum_value_index(list)


def q1(array):
    q1 = array[90:110, 70:84] #resizing array to cover just question 1
    male = q1[:10, :]
    female = q1[10:, :]
    return yes_no_q(male, female)


def q2(array):
    q2 = array[130:170, 50:135]
    q2 = cv2.resize(q2, (120, 60))
    age_18 = q2[:20, :20]
    age_1829 = q2[20:40, :20]
    age_3039 = q2[40:, :20]
    age_4049 = q2[:20, 100:]
    age_5059 = q2[20:40, 100:]
    age_60 = q2[40:, 100:]
    options = [age_18, age_1829, age_3039, age_4049, age_5059, age_60]
    return minimum_value_index(options)


def q3(array):
    q3 = array[200:265, 49:148]
    c_1 = q3[10:20, :9]
    c_2 = q3[30:40, :9]
    c_3 = q3[55:, :9]
    c_4 = q3[10:20, 90:]
    c_5 = q3[30:40, 90:]
    options = [c_1, c_2, c_3, c_4, c_5]
    return minimum_value_index(options)

def q4(array):
    q_4 = array[290:315, 77:87]
    q4y = np.sum(q_4[:12, :])
    q4n = np.sum(q_4[12:, :])
    return yes_no_q(q4y, q4n)


def q5(array):
    q5 = array[340:412, 65:78]
    c1 = q5[:12, :]
    c2 = q5[12:24, :]
    c3 = q5[24:36, :]
    c4 = q5[36:48, :]
    c5 = q5[48:60, :]
    c6 = q5[60:, :]
    opt = [c1, c2, c3, c4, c5, c6]
    return minimum_value_index(opt)

def q6(array):
    q6 = array[440:460, 78:88]
    y = q6[:10, :]
    n = q6[10:, :]
    return yes_no_q(y, n)


def q7(array):
    q7 = array[110:160, 260:390]
    c1 = q7[:15, :17]
    c2 = q7[15:30, :17]
    c3 = q7[30:45, :17]
    c4 = q7[:15, 33:50]
    c5 = q7[15:30, 33:50]
    c6 = q7[30:45, 33:50]
    c7 = q7[:15, 65:82]
    c8 = q7[15:30, 65:82]
    c9 = q7[30:45, 65:82]
    c10 = q7[:15, 98:115]
    options = [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10]
    return minimum_value_index(options)


def q8(array):
    q8 = array[200:240, 260:400]
    c1 = q8[:20, :15]
    c2 = q8[20:, :15]
    c3 = q8[:20, 60:75]
    c4 = q8[20:, 60:75]
    c5 = q8[:20, 120:135]
    c6 = q8[20:, 120:135]
    options = [c1, c2, c3, c4, c5, c6]
    return minimum_value_index(options)


def q9(array):
    q9 = array[285:315, 260:400]
    c1 = q9[:15, :17]
    c2 = q9[15:, :17]
    c3 = q9[:15, 60:77]
    c4 = q9[15:, 60:77]
    c5 = q9[:15, 110:127]
    c6 = q9[15:, 110:127]
    options = [c1, c2, c3, c4, c5, c6]
    return minimum_value_index(options)


def q10(array):
    q10 = array[355:465, 265:280]
    c1 = q10[:10, :]
    c2 = q10[10:20, :]
    c3 = q10[22:32, :]
    c4 = q10[33:43, :]
    c5 = q10[45:55, :]
    c6 = q10[55:65, :]
    c7 = q10[65:75, :]
    c8 = q10[77:87, :]
    c9 = q10[90:100, :]
    c10 = q10[100:110, :]
    options = [c1, c2, c3, c4, c5, c6, c7, c8, c9, c10]
    return all_that_apply(options)


def q11(array):
    answers_q11 = []
    row = 110
    column = 130
    while row != 190: 
        q11 = array[row:column, 225:465]
        c1 = q11[:, :60]
        c2 = q11[:, 60:120]
        c3 = q11[:, 120:180]
        c4 = q11[:, 180:]
        row += 20
        column += 20
        options = [c1, c2, c3, c4]
        answers_q11.append(minimum_value_index(options))
    return answers_q11


def q13(array):
    q13 = array[310:365, 55:70]
    c1 = q13[:11, :]
    c2 = q13[11:22, :]
    c3 = q13[22:33, :]
    c4 = q13[33:44, :]
    c5 = q13[44:, :]
    options = [c1, c2, c3, c4, c5]
    return minimum_value_index(options)
 

destination = str(today)+ '.xlsx'
batch_answer = []
for img in os.listdir(home):
    answers = []
    print(img)
    print(int(img[-5])% 2)
    file_num = int(img[-5])% 2
    if file_num != 0:
        img_array = cv2.imread(os.path.join(home,img), cv2.IMREAD_GRAYSCALE)
        IMG_SIZE = 500
        new_array = cv2.resize(img_array, (IMG_SIZE, IMG_SIZE))
        answers.append(q1(new_array))
        answers.append(q2(new_array))
        answers.append(q3(new_array))
        answers.append(q4(new_array))
        answers.append(q5(new_array))
        answers.append(q6(new_array))
        answers.append(q7(new_array))
        answers.append(q8(new_array))
        answers.append(q9(new_array))
        answers.append(q10(new_array))
        combined = answers
              
    elif file_num == 0:
        img_array = cv2.imread(os.path.join(home,img), cv2.IMREAD_GRAYSCALE)
        IMG_SIZE = 500
        new_array = cv2.resize(img_array, (IMG_SIZE, IMG_SIZE))
        answers.append(q11(new_array))
        answers.append(q13(new_array))
        total = combined + answers
        batch_answer.append(total)


os.chdir(output)
if os.path.isfile(output + destination):
    wb = openpyxl.load_workbook(destination)
else:
    wb = openpyxl.load_workbook('output.xlsx')
sheet = wb['Sheet1']
file_number = 1
for m in tqdm(batch_answer):
    print('total')
    print(m)
    count = 0
    file_number +=1
    for i in range(2, len(m)+2):
        if count <= len(m):
            sheet.cell(row=file_number, column=i).value = str(m[count])
            sheet.cell(row=file_number, column=1).value = str(today)
            count += 1
            wb.save(destination)
